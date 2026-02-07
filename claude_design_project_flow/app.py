"""Application Infrastructure Onboarding Portal — FastAPI + Jinja2 + Tailwind CSS + vis-network + Chart.js"""

from fastapi import FastAPI, Request, HTTPException, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import JSONResponse, StreamingResponse
import sqlite3
import json
import csv
import io
from datetime import datetime, date, timedelta
from contextlib import contextmanager
import os
import uvicorn

from init_db import DB_PATH, init_database, create_project_from_template

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = FastAPI(title="Application Infrastructure Onboarding Portal")
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))


# ── DB helpers ──────────────────────────────────────────────────────────────

@contextmanager
def get_db():
    conn = sqlite3.connect(os.path.join(BASE_DIR, DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
    finally:
        conn.close()


def rows(cursor_result):
    return [dict(r) for r in cursor_result]


def row(cursor_result):
    return dict(cursor_result) if cursor_result else None


# ── Helpers ─────────────────────────────────────────────────────────────────

def rag_status(project):
    if not project.get("start_date") or not project.get("target_completion"):
        return "gray"
    start = datetime.strptime(project["start_date"], "%Y-%m-%d").date()
    target = datetime.strptime(project["target_completion"], "%Y-%m-%d").date()
    today = date.today()
    total_days = max((target - start).days, 1)
    elapsed = (today - start).days
    expected_pct = min(100, max(0, elapsed / total_days * 100))
    total = project.get("total_tasks") or 1
    actual_pct = (project.get("completed_tasks", 0) / total) * 100
    diff = actual_pct - expected_pct
    if diff >= -10:
        return "green"
    if diff >= -25:
        return "amber"
    return "red"


RAG_COLORS = {"green": "#22c55e", "amber": "#3b82f6", "red": "#ef4444", "gray": "#94a3b8"}
RAG_LABELS = {"green": "On Track", "amber": "In Progress", "red": "Delayed", "gray": "No Data"}

templates.env.globals.update(rag_status=rag_status, RAG_COLORS=RAG_COLORS, RAG_LABELS=RAG_LABELS)


def base_context(request: Request, page: str) -> dict:
    with get_db() as db:
        sidebar_projects = rows(db.execute(
            "SELECT id, project_name, domain, environment FROM projects ORDER BY domain, environment, project_name"
        ).fetchall())
        sidebar_domains = rows(db.execute(
            "SELECT code, display_name FROM domains WHERE is_active=1 ORDER BY code"
        ).fetchall())
    return {"request": request, "page": page, "sidebar_projects": sidebar_projects, "sidebar_domains": sidebar_domains}


# ══════════════════════════════════════════════════════════════════════════════
# PAGE ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/")
async def dashboard(request: Request):
    view = request.query_params.get("view", "portfolio")
    filter_domain = request.query_params.get("domain", "")
    filter_status = request.query_params.get("status", "")
    ctx = base_context(request, "dashboard")
    with get_db() as db:
        # Base query
        query = """
            SELECT p.*,
                (SELECT COUNT(*) FROM tasks t JOIN phases ph ON t.phase_id=ph.id WHERE ph.project_id=p.id) as total_tasks,
                (SELECT COUNT(*) FROM tasks t JOIN phases ph ON t.phase_id=ph.id WHERE ph.project_id=p.id AND t.status='completed') as completed_tasks,
                (SELECT COUNT(*) FROM phases WHERE project_id=p.id AND status='completed') as completed_phases,
                (SELECT COUNT(*) FROM phases WHERE project_id=p.id) as total_phases,
                (SELECT COUNT(*) FROM phases WHERE project_id=p.id AND status='in_progress') as active_phases
            FROM projects p
        """
        params = []
        if filter_domain:
            query += " WHERE p.domain=?"
            params.append(filter_domain)
        query += " ORDER BY p.created_at DESC"
        all_projects = rows(db.execute(query, params).fetchall())

        # Filter by RAG status if needed
        if filter_status:
            projects = [p for p in all_projects if rag_status(p) == filter_status]
        else:
            projects = all_projects

        domains = rows(db.execute("""
            SELECT domain, COUNT(*) as count,
                SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) as completed
            FROM projects GROUP BY domain
        """).fetchall())

        # Execution view stats
        task_stats = db.execute("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN status='in_progress' THEN 1 ELSE 0 END) as in_progress,
                SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END) as pending,
                SUM(CASE WHEN is_blocked=1 THEN 1 ELSE 0 END) as blocked,
                SUM(CASE WHEN ownership='Platform Team' THEN 1 ELSE 0 END) as platform,
                SUM(CASE WHEN ownership='App Team' THEN 1 ELSE 0 END) as app,
                SUM(CASE WHEN status='completed' AND date(completed_at)=date('now') THEN 1 ELSE 0 END) as today
            FROM tasks
        """).fetchone()

        # Management view stats
        blocked_items = rows(db.execute("""
            SELECT t.task_name, p.project_name
            FROM tasks t
            JOIN phases ph ON t.phase_id=ph.id
            JOIN projects p ON ph.project_id=p.id
            WHERE t.is_blocked=1 LIMIT 10
        """).fetchall())

        overdue_phases = rows(db.execute("""
            SELECT ph.phase_number, ph.phase_name, p.project_name,
                julianday('now') - julianday(ph.target_end) as days_overdue
            FROM phases ph
            JOIN projects p ON ph.project_id=p.id
            WHERE ph.status != 'completed' AND ph.target_end < date('now')
            ORDER BY days_overdue DESC LIMIT 10
        """).fetchall())

        avg_variance = db.execute("""
            SELECT COALESCE(AVG(variance_days), 0) FROM phase_deadlines
        """).fetchone()[0]

        # Count of templates for getting started wizard
        templates_count = db.execute("SELECT COUNT(*) FROM templates").fetchone()[0]

    rag_counts = {"green": 0, "amber": 0, "red": 0, "gray": 0}
    projects_by_domain = {}  # domain -> {green: count, amber: count, red: count}
    for p in projects:
        status = rag_status(p)
        rag_counts[status] += 1
        domain = p.get("domain", "Unknown")
        if domain not in projects_by_domain:
            projects_by_domain[domain] = {"green": 0, "amber": 0, "red": 0, "gray": 0}
        projects_by_domain[domain][status] += 1

    ctx.update(
        view=view,
        projects=projects,
        domains=domains,
        total_projects=len(projects),
        rag_counts=rag_counts,
        projects_by_domain=projects_by_domain,
        # Execution view
        total_tasks=task_stats[0] or 0,
        completed_tasks=task_stats[1] or 0,
        in_progress_tasks=task_stats[2] or 0,
        pending_tasks=task_stats[3] or 0,
        blocked_tasks=task_stats[4] or 0,
        platform_tasks=task_stats[5] or 0,
        app_tasks=task_stats[6] or 0,
        completed_today=task_stats[7] or 0,
        # Management view
        blocked_items=blocked_items,
        overdue_phases=overdue_phases,
        avg_variance=round(avg_variance, 1) if avg_variance else 0,
        escalations=len(blocked_items) + len(overdue_phases),
        # Getting started wizard
        templates_count=templates_count
    )
    return templates.TemplateResponse("dashboard.html", ctx)


@app.get("/domain/{domain_code}")
async def domain_dashboard(request: Request, domain_code: str):
    """Domain-specific dashboard showing all projects for that domain."""
    ctx = base_context(request, "domain")
    ctx["active_domain"] = domain_code

    with get_db() as db:
        # Get domain info
        domain_info = db.execute(
            "SELECT * FROM domains WHERE code=?", (domain_code,)
        ).fetchone()
        if not domain_info:
            domain_info = {"code": domain_code, "display_name": domain_code}
        else:
            domain_info = row(domain_info)

        # Get projects for this domain
        projects = rows(db.execute("""
            SELECT p.*,
                (SELECT COUNT(*) FROM tasks t JOIN phases ph ON t.phase_id=ph.id WHERE ph.project_id=p.id) as total_tasks,
                (SELECT COUNT(*) FROM tasks t JOIN phases ph ON t.phase_id=ph.id WHERE ph.project_id=p.id AND t.status='completed') as completed_tasks,
                (SELECT COUNT(*) FROM phases WHERE project_id=p.id AND status='completed') as completed_phases,
                (SELECT COUNT(*) FROM phases WHERE project_id=p.id) as total_phases,
                (SELECT COUNT(*) FROM phases WHERE project_id=p.id AND status='in_progress') as active_phases
            FROM projects p WHERE p.domain=? ORDER BY p.environment, p.created_at DESC
        """, (domain_code,)).fetchall())

        # Get environments with project counts
        envs = rows(db.execute("""
            SELECT e.code, e.display_name, e.sort_order,
                COALESCE((SELECT COUNT(*) FROM projects WHERE domain=? AND environment=e.code), 0) as project_count,
                COALESCE((SELECT SUM(CASE WHEN t.status='completed' THEN 1 ELSE 0 END)
                         FROM tasks t JOIN phases ph ON t.phase_id=ph.id
                         JOIN projects p ON ph.project_id=p.id
                         WHERE p.domain=? AND p.environment=e.code), 0) as completed_tasks,
                COALESCE((SELECT COUNT(*)
                         FROM tasks t JOIN phases ph ON t.phase_id=ph.id
                         JOIN projects p ON ph.project_id=p.id
                         WHERE p.domain=? AND p.environment=e.code), 0) as total_tasks
            FROM environments e ORDER BY e.sort_order
        """, (domain_code, domain_code, domain_code)).fetchall())

        # Task stats for domain
        task_stats = db.execute("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN t.status='completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN t.status='in_progress' THEN 1 ELSE 0 END) as in_progress,
                SUM(CASE WHEN t.status='pending' THEN 1 ELSE 0 END) as pending
            FROM tasks t
            JOIN phases ph ON t.phase_id=ph.id
            JOIN projects p ON ph.project_id=p.id
            WHERE p.domain=?
        """, (domain_code,)).fetchone()

    rag_counts = {"green": 0, "amber": 0, "red": 0, "gray": 0}
    for p in projects:
        rag_counts[rag_status(p)] += 1

    ctx.update(
        domain=domain_info,
        domain_code=domain_code,
        projects=projects,
        environments=envs,
        rag_counts=rag_counts,
        total_projects=len(projects),
        total_tasks=task_stats[0] or 0,
        completed_tasks=task_stats[1] or 0,
        in_progress_tasks=task_stats[2] or 0,
        pending_tasks=task_stats[3] or 0,
        progress_pct=round((task_stats[1] or 0) / max(task_stats[0] or 1, 1) * 100)
    )
    return templates.TemplateResponse("domain_dashboard.html", ctx)


@app.get("/project/{project_id}")
async def project_detail(request: Request, project_id: int):
    ctx = base_context(request, "project")
    with get_db() as db:
        project = row(db.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone())
        if not project:
            raise HTTPException(404, "Project not found")

        phases = rows(db.execute("""
            SELECT ph.*,
                (SELECT COUNT(*) FROM tasks WHERE phase_id=ph.id) as total_tasks,
                (SELECT COUNT(*) FROM tasks WHERE phase_id=ph.id AND status='completed') as completed_tasks,
                (SELECT COUNT(*) FROM tasks WHERE phase_id=ph.id AND status='in_progress') as active_tasks
            FROM phases ph WHERE ph.project_id=? ORDER BY ph.phase_number
        """, (project_id,)).fetchall())

        deps = rows(db.execute("""
            SELECT pd.phase_id, pd.depends_on_phase_id,
                p1.phase_number as phase_num, p2.phase_number as dep_num,
                p2.status as dep_status
            FROM phase_dependencies pd
            JOIN phases p1 ON pd.phase_id=p1.id
            JOIN phases p2 ON pd.depends_on_phase_id=p2.id
            WHERE pd.project_id=?
        """, (project_id,)).fetchall())

        tasks_by_phase = {}
        for ph in phases:
            tasks_by_phase[ph["id"]] = rows(db.execute("""
                SELECT t.*,
                    (SELECT GROUP_CONCAT(action_label || '|' || action_url, ';;') FROM task_actions WHERE task_id=t.id) as actions_raw
                FROM tasks t WHERE t.phase_id=? ORDER BY t.task_order
            """, (ph["id"],)).fetchall())

        dep_map = {}
        for d in deps:
            dep_map.setdefault(d["phase_id"], []).append(d)
        phase_status = {ph["id"]: ph["status"] for ph in phases}
        for ph in phases:
            prereqs = dep_map.get(ph["id"], [])
            ph["is_locked"] = any(phase_status.get(d["depends_on_phase_id"]) != "completed" for d in prereqs)
            ph["prerequisites"] = prereqs
            ph["progress"] = int(ph["completed_tasks"] / ph["total_tasks"] * 100) if ph["total_tasks"] else 0

        total_tasks = sum(ph["total_tasks"] for ph in phases)
        completed_tasks = sum(ph["completed_tasks"] for ph in phases)

    ctx.update(project=project, phases=phases, deps=deps,
               tasks_by_phase=tasks_by_phase,
               total_tasks=total_tasks, completed_tasks=completed_tasks)
    return templates.TemplateResponse("project.html", ctx)


@app.get("/tasks")
async def task_list(request: Request, project_id: int = 0, phase: int = 0,
                    status: str = "", ownership: str = ""):
    ctx = base_context(request, "tasks")
    with get_db() as db:
        projects = rows(db.execute(
            "SELECT id, project_name, domain FROM projects ORDER BY project_name"
        ).fetchall())

        query = """
            SELECT t.*, ph.phase_number, ph.phase_name, p.project_name, p.domain, p.id as project_id,
                (SELECT GROUP_CONCAT(action_label || '|' || action_url, ';;') FROM task_actions WHERE task_id=t.id) as actions_raw
            FROM tasks t
            JOIN phases ph ON t.phase_id=ph.id
            JOIN projects p ON ph.project_id=p.id WHERE 1=1
        """
        params = []
        if project_id:
            query += " AND p.id=?"
            params.append(project_id)
        if phase:
            query += " AND ph.phase_number=?"
            params.append(phase)
        if status:
            query += " AND t.status=?"
            params.append(status)
        if ownership:
            query += " AND t.ownership=?"
            params.append(ownership)
        query += " ORDER BY p.project_name, ph.phase_number, t.task_order"

        tasks = rows(db.execute(query, params).fetchall())

        # Get selected project if project_id is set
        selected_project = None
        if project_id:
            selected_project = row(db.execute(
                "SELECT id, project_name, domain FROM projects WHERE id=?", (project_id,)
            ).fetchone())
            # Stats for selected project only
            stats = row(db.execute("""
                SELECT COUNT(*) as total,
                    SUM(CASE WHEN t.status='completed' THEN 1 ELSE 0 END) as completed,
                    SUM(CASE WHEN t.status='in_progress' THEN 1 ELSE 0 END) as in_progress,
                    SUM(CASE WHEN t.status='pending' THEN 1 ELSE 0 END) as pending
                FROM tasks t
                JOIN phases ph ON t.phase_id=ph.id
                WHERE ph.project_id=?
            """, (project_id,)).fetchone())
            project_stats = []
        else:
            # Global stats
            stats = row(db.execute("""
                SELECT COUNT(*) as total,
                    SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) as completed,
                    SUM(CASE WHEN status='in_progress' THEN 1 ELSE 0 END) as in_progress,
                    SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END) as pending
                FROM tasks
            """).fetchone())
            # Per-project breakdown
            project_stats = rows(db.execute("""
                SELECT p.id, p.project_name, p.domain,
                    COUNT(t.id) as total,
                    SUM(CASE WHEN t.status='completed' THEN 1 ELSE 0 END) as completed,
                    SUM(CASE WHEN t.status='in_progress' THEN 1 ELSE 0 END) as in_progress,
                    SUM(CASE WHEN t.status='pending' THEN 1 ELSE 0 END) as pending
                FROM projects p
                LEFT JOIN phases ph ON ph.project_id=p.id
                LEFT JOIN tasks t ON t.phase_id=ph.id
                GROUP BY p.id
                ORDER BY p.project_name
            """).fetchall())

    ctx.update(tasks=tasks, all_projects=projects, stats=stats,
               selected_project=selected_project, project_stats=project_stats,
               filters={"project_id": project_id, "phase": phase,
                        "status": status, "ownership": ownership})
    return templates.TemplateResponse("tasks.html", ctx)


# ── Template Pages ──────────────────────────────────────────────────────────

@app.get("/templates")
async def template_list(request: Request):
    ctx = base_context(request, "templates")
    with get_db() as db:
        tpls = rows(db.execute("""
            SELECT t.*,
                (SELECT COUNT(*) FROM template_phases WHERE template_id=t.id) as phase_count,
                (SELECT COUNT(*) FROM template_tasks tt
                 JOIN template_phases tp ON tt.template_phase_id=tp.id
                 WHERE tp.template_id=t.id) as task_count,
                (SELECT COUNT(*) FROM template_phase_dependencies WHERE template_id=t.id) as dep_count
            FROM templates t ORDER BY t.is_default DESC, t.created_at DESC
        """).fetchall())
    ctx.update(tpls=tpls)
    return templates.TemplateResponse("templates_list.html", ctx)


@app.get("/templates/new")
async def template_new(request: Request):
    ctx = base_context(request, "templates")
    ctx.update(template=None, template_phases=[], tasks_by_phase={}, deps=[], is_new=True)
    return templates.TemplateResponse("template_editor.html", ctx)


@app.get("/templates/{template_id}")
async def template_editor(request: Request, template_id: int):
    ctx = base_context(request, "templates")
    with get_db() as db:
        tpl = row(db.execute("SELECT * FROM templates WHERE id=?", (template_id,)).fetchone())
        if not tpl:
            raise HTTPException(404, "Template not found")

        t_phases = rows(db.execute("""
            SELECT tp.*,
                (SELECT COUNT(*) FROM template_tasks WHERE template_phase_id=tp.id) as task_count
            FROM template_phases tp WHERE tp.template_id=? ORDER BY tp.phase_number
        """, (template_id,)).fetchall())

        tasks_by_phase = {}
        for tp in t_phases:
            tasks_by_phase[tp["id"]] = rows(db.execute("""
                SELECT tt.*,
                    (SELECT GROUP_CONCAT(action_label || '|' || action_url, ';;')
                     FROM template_task_actions WHERE template_task_id=tt.id) as actions_raw
                FROM template_tasks tt WHERE tt.template_phase_id=? ORDER BY tt.task_order
            """, (tp["id"],)).fetchall())

        dep_list = rows(db.execute(
            "SELECT * FROM template_phase_dependencies WHERE template_id=?", (template_id,)
        ).fetchall())

    ctx.update(template=tpl, template_phases=t_phases,
               tasks_by_phase=tasks_by_phase, deps=dep_list, is_new=False)
    return templates.TemplateResponse("template_editor.html", ctx)


@app.get("/projects/new")
async def new_project_page(request: Request):
    ctx = base_context(request, "new_project")
    preselected_domain = request.query_params.get("domain", "")
    preselected_env = request.query_params.get("env", "")
    with get_db() as db:
        tpls = rows(db.execute("""
            SELECT t.id, t.template_name, t.description, t.is_default,
                (SELECT COUNT(*) FROM template_phases WHERE template_id=t.id) as phase_count,
                (SELECT COUNT(*) FROM template_tasks tt
                 JOIN template_phases tp ON tt.template_phase_id=tp.id
                 WHERE tp.template_id=t.id) as task_count
            FROM templates t ORDER BY t.is_default DESC, t.template_name
        """).fetchall())
        domain_list = rows(db.execute(
            "SELECT * FROM domains WHERE is_active=1 ORDER BY code"
        ).fetchall())
        env_list = rows(db.execute(
            "SELECT * FROM environments ORDER BY sort_order, code"
        ).fetchall())
    ctx.update(
        templates_list=tpls,
        domains=domain_list,
        environments=env_list,
        preselected_domain=preselected_domain,
        preselected_env=preselected_env
    )
    return templates.TemplateResponse("new_project.html", ctx)


# ══════════════════════════════════════════════════════════════════════════════
# API ENDPOINTS — Projects
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/project/{project_id}/dag")
async def api_dag(project_id: int):
    with get_db() as db:
        phases = rows(db.execute("""
            SELECT ph.*,
                (SELECT COUNT(*) FROM tasks WHERE phase_id=ph.id) as total_tasks,
                (SELECT COUNT(*) FROM tasks WHERE phase_id=ph.id AND status='completed') as completed_tasks
            FROM phases ph WHERE ph.project_id=? ORDER BY ph.phase_number
        """, (project_id,)).fetchall())

        deps = rows(db.execute("""
            SELECT p1.phase_number as from_phase, p2.phase_number as to_phase,
                   p1.status as from_status
            FROM phase_dependencies pd
            JOIN phases p1 ON pd.depends_on_phase_id=p1.id
            JOIN phases p2 ON pd.phase_id=p2.id
            WHERE pd.project_id=?
        """, (project_id,)).fetchall())

    status_colors = {"completed": "#22c55e", "in_progress": "#3b82f6", "pending": "#94a3b8"}
    status_borders = {"completed": "#16a34a", "in_progress": "#2563eb", "pending": "#64748b"}
    short = {
        "Prerequisites": "Prerequisites",
        "Infrastructure and Pipeline Setup": "Infra & Pipeline",
        "Application Deployment": "App Deployment",
        "Data Migration": "Data Migration", "Upgrade": "Upgrade",
        "Logging and Monitoring": "Logging & Monitoring",
        "Performance": "Performance", "Resiliency": "Resiliency", "Security": "Security",
    }

    nodes = []
    for ph in phases:
        pct = int(ph["completed_tasks"] / ph["total_tasks"] * 100) if ph["total_tasks"] else 0
        label = short.get(ph["phase_name"], ph["phase_name"])
        nodes.append({
            "id": ph["phase_number"],
            "label": f"P{ph['phase_number']}: {label}\n{pct}% ({ph['completed_tasks']}/{ph['total_tasks']})",
            "color": {"background": status_colors.get(ph["status"], "#94a3b8"),
                      "border": status_borders.get(ph["status"], "#64748b"),
                      "highlight": {"background": status_colors.get(ph["status"]), "border": "#1e293b"}},
            "font": {"color": "#ffffff", "size": 13, "face": "Inter, system-ui, sans-serif", "multi": "md"},
            "shape": "box", "borderWidth": 2, "borderWidthSelected": 3,
            "shadow": {"enabled": True, "size": 8, "x": 2, "y": 2},
            "margin": {"top": 12, "bottom": 12, "left": 16, "right": 16},
            "widthConstraint": {"minimum": 160, "maximum": 200},
            "phase_id": ph["id"], "status": ph["status"], "progress": pct,
        })

    edges = []
    for d in deps:
        color = "#22c55e" if d["from_status"] == "completed" else "#cbd5e1"
        edges.append({
            "from": d["from_phase"], "to": d["to_phase"],
            "arrows": {"to": {"enabled": True, "scaleFactor": 0.8}},
            "color": {"color": color, "highlight": "#1e293b"},
            "width": 2.5,
            "smooth": {"type": "cubicBezier", "forceDirection": "horizontal", "roundness": 0.4},
        })
    return {"nodes": nodes, "edges": edges}


@app.get("/api/project/{project_id}/stats")
async def api_stats(project_id: int):
    with get_db() as db:
        task_stats = rows(db.execute("""
            SELECT status, COUNT(*) as count FROM tasks
            WHERE phase_id IN (SELECT id FROM phases WHERE project_id=?)
            GROUP BY status
        """, (project_id,)).fetchall())
        phase_stats = rows(db.execute("""
            SELECT ph.phase_name, ph.phase_number,
                COUNT(*) as total,
                SUM(CASE WHEN t.status='completed' THEN 1 ELSE 0 END) as completed
            FROM tasks t JOIN phases ph ON t.phase_id=ph.id
            WHERE ph.project_id=?
            GROUP BY ph.id ORDER BY ph.phase_number
        """, (project_id,)).fetchall())
        ownership_stats = rows(db.execute("""
            SELECT ownership, COUNT(*) as total,
                SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) as completed
            FROM tasks WHERE phase_id IN (SELECT id FROM phases WHERE project_id=?)
            GROUP BY ownership
        """, (project_id,)).fetchall())
        timeline = rows(db.execute("""
            SELECT phase_number, phase_name, target_start, target_end,
                   started_at, completed_at, status
            FROM phases WHERE project_id=? ORDER BY phase_number
        """, (project_id,)).fetchall())
    return {"task_stats": {s["status"]: s["count"] for s in task_stats},
            "phase_stats": phase_stats, "ownership_stats": ownership_stats, "timeline": timeline}


@app.post("/api/task/{task_id}/status")
async def update_task_status(task_id: int, request: Request):
    body = await request.json()
    new_status = body.get("status")
    if new_status not in ("pending", "in_progress", "completed"):
        raise HTTPException(400, "Invalid status")
    with get_db() as db:
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if new_status == "in_progress":
            db.execute("UPDATE tasks SET status=?, started_at=COALESCE(started_at,?) WHERE id=?",
                       (new_status, now_str, task_id))
        elif new_status == "completed":
            db.execute("UPDATE tasks SET status=?, completed_at=? WHERE id=?",
                       (new_status, now_str, task_id))
        else:
            db.execute("UPDATE tasks SET status=?, started_at=NULL, completed_at=NULL WHERE id=?",
                       (new_status, task_id))
        t = row(db.execute("SELECT phase_id FROM tasks WHERE id=?", (task_id,)).fetchone())
        if t:
            pid = t["phase_id"]
            total = db.execute("SELECT COUNT(*) FROM tasks WHERE phase_id=?", (pid,)).fetchone()[0]
            done = db.execute("SELECT COUNT(*) FROM tasks WHERE phase_id=? AND status='completed'", (pid,)).fetchone()[0]
            prog = db.execute("SELECT COUNT(*) FROM tasks WHERE phase_id=? AND status='in_progress'", (pid,)).fetchone()[0]
            if done == total and total > 0:
                db.execute("UPDATE phases SET status='completed', completed_at=? WHERE id=?", (now_str, pid))
            elif prog > 0 or done > 0:
                db.execute("UPDATE phases SET status='in_progress', started_at=COALESCE(started_at,?) WHERE id=?", (now_str, pid))
            else:
                db.execute("UPDATE phases SET status='pending' WHERE id=?", (pid,))
        db.commit()
    return {"ok": True}


@app.post("/api/task/create")
async def create_task(request: Request):
    body = await request.json()
    with get_db() as db:
        mx = db.execute("SELECT COALESCE(MAX(task_order),0) FROM tasks WHERE phase_id=?",
                        (body["phase_id"],)).fetchone()[0]
        db.execute("""INSERT INTO tasks (phase_id, task_order, task_name, category, description,
                      ownership, estimated_minutes, priority, status) VALUES (?,?,?,?,?,?,?,?,?)""",
                   (body["phase_id"], mx + 1, body["task_name"],
                    body.get("category", "General"), body.get("description", body["task_name"]),
                    body.get("ownership", "Platform Team"),
                    body.get("estimated_minutes", 30), body.get("priority", "medium"), "pending"))
        db.commit()
        new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return {"ok": True, "task_id": new_id}


@app.put("/api/task/{task_id}")
async def update_task(task_id: int, request: Request):
    """Update task details."""
    body = await request.json()
    with get_db() as db:
        db.execute("""UPDATE tasks SET
                      task_name=COALESCE(?, task_name),
                      category=COALESCE(?, category),
                      ownership=COALESCE(?, ownership),
                      jira_reference=?, sample_reference=?, notes=?,
                      description=COALESCE(?, description)
                      WHERE id=?""",
                   (body.get("task_name"), body.get("category"), body.get("ownership"),
                    body.get("jira_reference"), body.get("sample_reference"), body.get("notes"),
                    body.get("description"), task_id))
        db.commit()
    return {"ok": True}


@app.delete("/api/task/{task_id}")
async def delete_task(task_id: int):
    """Delete a task."""
    with get_db() as db:
        db.execute("DELETE FROM tasks WHERE id=?", (task_id,))
        db.commit()
    return {"ok": True}


@app.post("/api/phase/create")
async def create_phase(request: Request):
    body = await request.json()
    with get_db() as db:
        db.execute("""INSERT INTO phases (project_id, phase_number, phase_name, status,
                      target_start, target_end) VALUES (?,?,?,?,?,?)""",
                   (body["project_id"], body["phase_number"], body["phase_name"],
                    "pending", body.get("target_start"), body.get("target_end")))
        phase_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        for dep_id in body.get("depends_on", []):
            db.execute("INSERT INTO phase_dependencies (project_id, phase_id, depends_on_phase_id) VALUES (?,?,?)",
                       (body["project_id"], phase_id, dep_id))
        db.commit()
    return {"ok": True, "phase_id": phase_id}


@app.put("/api/phase/{phase_id}")
async def update_phase(phase_id: int, request: Request):
    """Update phase details."""
    body = await request.json()
    with get_db() as db:
        db.execute("""UPDATE phases SET
                      phase_name=COALESCE(?, phase_name),
                      phase_number=COALESCE(?, phase_number),
                      status=COALESCE(?, status),
                      target_start=COALESCE(?, target_start),
                      target_end=COALESCE(?, target_end)
                      WHERE id=?""",
                   (body.get("phase_name"), body.get("phase_number"), body.get("status"),
                    body.get("target_start"), body.get("target_end"), phase_id))
        db.commit()
    return {"ok": True}


@app.delete("/api/phase/{phase_id}")
async def delete_phase(phase_id: int):
    """Delete a phase and all its tasks."""
    with get_db() as db:
        db.execute("DELETE FROM phases WHERE id=?", (phase_id,))
        db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# API ENDPOINTS — Domains & Environments
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/api/domain")
async def api_create_domain(request: Request):
    body = await request.json()
    with get_db() as db:
        # Check if domain already exists
        existing = db.execute("SELECT id FROM domains WHERE code=?", (body["code"],)).fetchone()
        if existing:
            raise HTTPException(400, "Domain already exists")
        db.execute("INSERT INTO domains (code, display_name, description) VALUES (?,?,?)",
                   (body["code"], body["display_name"], body.get("description", "")))
        db.commit()
        did = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return {"ok": True, "domain_id": did}


@app.put("/api/domain/{domain_id}")
async def api_update_domain(domain_id: int, request: Request):
    body = await request.json()
    with get_db() as db:
        db.execute("UPDATE domains SET display_name=?, description=?, is_active=? WHERE id=?",
                   (body.get("display_name"), body.get("description"), body.get("is_active", 1), domain_id))
        db.commit()
    return {"ok": True}


@app.delete("/api/domain/{domain_id}")
async def api_delete_domain(domain_id: int):
    with get_db() as db:
        db.execute("DELETE FROM domains WHERE id=?", (domain_id,))
        db.commit()
    return {"ok": True}


@app.post("/api/environment")
async def api_create_environment(request: Request):
    body = await request.json()
    with get_db() as db:
        existing = db.execute("SELECT id FROM environments WHERE code=?", (body["code"],)).fetchone()
        if existing:
            raise HTTPException(400, "Environment already exists")
        db.execute("INSERT INTO environments (code, display_name, sort_order) VALUES (?,?,?)",
                   (body["code"], body["display_name"], body.get("sort_order", 0)))
        db.commit()
        eid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return {"ok": True, "environment_id": eid}


@app.put("/api/environment/{env_id}")
async def api_update_environment(env_id: int, request: Request):
    body = await request.json()
    with get_db() as db:
        db.execute("UPDATE environments SET display_name=?, sort_order=? WHERE id=?",
                   (body.get("display_name"), body.get("sort_order", 0), env_id))
        db.commit()
    return {"ok": True}


@app.delete("/api/environment/{env_id}")
async def api_delete_environment(env_id: int):
    with get_db() as db:
        db.execute("DELETE FROM environments WHERE id=?", (env_id,))
        db.commit()
    return {"ok": True}


@app.get("/api/domains")
async def api_list_domains():
    with get_db() as db:
        result = rows(db.execute("SELECT * FROM domains ORDER BY code").fetchall())
    return result


@app.get("/api/environments")
async def api_list_environments():
    with get_db() as db:
        result = rows(db.execute("SELECT * FROM environments ORDER BY sort_order, code").fetchall())
    return result


# ══════════════════════════════════════════════════════════════════════════════
# API ENDPOINTS — Templates
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/api/template/create")
async def api_create_template(request: Request):
    body = await request.json()
    with get_db() as db:
        db.execute("INSERT INTO templates (template_name, description, version) VALUES (?,?,?)",
                   (body["template_name"], body.get("description", ""), body.get("version", "1.0")))
        db.commit()
        tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return {"ok": True, "template_id": tid}


@app.put("/api/template/{template_id}")
async def api_update_template(template_id: int, request: Request):
    body = await request.json()
    with get_db() as db:
        db.execute("""UPDATE templates SET template_name=?, description=?, version=?,
                      updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                   (body["template_name"], body.get("description", ""),
                    body.get("version", "1.0"), template_id))
        db.commit()
    return {"ok": True}


@app.delete("/api/template/{template_id}")
async def api_delete_template(template_id: int):
    with get_db() as db:
        tpl = row(db.execute("SELECT is_default FROM templates WHERE id=?", (template_id,)).fetchone())
        if not tpl:
            raise HTTPException(404, "Template not found")
        if tpl["is_default"]:
            raise HTTPException(400, "Cannot delete the default template")
        count = db.execute("SELECT COUNT(*) FROM projects WHERE template_id=?", (template_id,)).fetchone()[0]
        if count > 0:
            raise HTTPException(400, f"Cannot delete: {count} project(s) use this template")
        db.execute("DELETE FROM templates WHERE id=?", (template_id,))
        db.commit()
    return {"ok": True}


@app.post("/api/template/{template_id}/clone")
async def api_clone_template(template_id: int, request: Request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    new_name = body.get("template_name", "")
    with get_db() as db:
        src = row(db.execute("SELECT * FROM templates WHERE id=?", (template_id,)).fetchone())
        if not src:
            raise HTTPException(404, "Source template not found")
        if not new_name:
            new_name = src["template_name"] + " (Copy)"

        db.execute("INSERT INTO templates (template_name, description, version) VALUES (?,?,?)",
                   (new_name, src["description"], src["version"]))
        new_tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        for sp in db.execute("SELECT * FROM template_phases WHERE template_id=? ORDER BY phase_number",
                             (template_id,)).fetchall():
            db.execute("INSERT INTO template_phases (template_id, phase_number, phase_name, duration_days) VALUES (?,?,?,?)",
                       (new_tid, sp["phase_number"], sp["phase_name"], sp["duration_days"]))
            new_tp_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

            for st in db.execute("SELECT * FROM template_tasks WHERE template_phase_id=? ORDER BY task_order",
                                 (sp["id"],)).fetchall():
                db.execute("""INSERT INTO template_tasks (template_phase_id, task_order, task_name, category,
                              description, ownership, instructions, estimated_minutes, priority,
                              automation_type, automation_config) VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                           (new_tp_id, st["task_order"], st["task_name"], st["category"],
                            st["description"], st["ownership"], st["instructions"],
                            st["estimated_minutes"], st["priority"],
                            st["automation_type"], st["automation_config"]))
                new_tt_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

                for sa in db.execute("SELECT * FROM template_task_actions WHERE template_task_id=?",
                                     (st["id"],)).fetchall():
                    db.execute("INSERT INTO template_task_actions VALUES (NULL,?,?,?,?)",
                               (new_tt_id, sa["action_type"], sa["action_label"], sa["action_url"]))

        for sd in db.execute("SELECT * FROM template_phase_dependencies WHERE template_id=?",
                             (template_id,)).fetchall():
            db.execute("""INSERT INTO template_phase_dependencies
                          (template_id, phase_number, depends_on_phase_number) VALUES (?,?,?)""",
                       (new_tid, sd["phase_number"], sd["depends_on_phase_number"]))
        db.commit()
        return {"ok": True, "template_id": new_tid, "template_name": new_name}


@app.get("/api/template/{template_id}/export")
async def api_export_template(template_id: int):
    with get_db() as db:
        tpl = row(db.execute("SELECT * FROM templates WHERE id=?", (template_id,)).fetchone())
        if not tpl:
            raise HTTPException(404)
        phase_list = []
        for tp in db.execute("SELECT * FROM template_phases WHERE template_id=? ORDER BY phase_number",
                             (template_id,)).fetchall():
            task_list = []
            for tt in db.execute("SELECT * FROM template_tasks WHERE template_phase_id=? ORDER BY task_order",
                                 (tp["id"],)).fetchall():
                actions = [{"action_type": a["action_type"], "action_label": a["action_label"],
                            "action_url": a["action_url"]}
                           for a in db.execute("SELECT * FROM template_task_actions WHERE template_task_id=?",
                                               (tt["id"],)).fetchall()]
                task_list.append({
                    "task_order": tt["task_order"], "task_name": tt["task_name"],
                    "category": tt["category"], "description": tt["description"],
                    "ownership": tt["ownership"], "instructions": tt["instructions"],
                    "estimated_minutes": tt["estimated_minutes"], "priority": tt["priority"],
                    "automation_type": tt["automation_type"],
                    "automation_config": json.loads(tt["automation_config"] or "{}"),
                    "actions": actions,
                })
            phase_list.append({
                "phase_number": tp["phase_number"], "phase_name": tp["phase_name"],
                "duration_days": tp["duration_days"], "tasks": task_list,
            })
        dep_list = [{"phase_number": d["phase_number"], "depends_on_phase_number": d["depends_on_phase_number"]}
                    for d in db.execute("SELECT * FROM template_phase_dependencies WHERE template_id=?",
                                        (template_id,)).fetchall()]
    export = {
        "template_name": tpl["template_name"], "description": tpl["description"],
        "version": tpl["version"], "phases": phase_list, "dependencies": dep_list,
        "exported_at": datetime.now().isoformat(), "format_version": "1.0",
    }
    return JSONResponse(content=export, headers={
        "Content-Disposition": f'attachment; filename="{tpl["template_name"].replace(" ", "_")}.json"'
    })


@app.post("/api/template/import")
async def api_import_template(request: Request):
    body = await request.json()
    with get_db() as db:
        name = body.get("template_name", "Imported Template")
        if db.execute("SELECT id FROM templates WHERE template_name=?", (name,)).fetchone():
            name += f" (imported {datetime.now().strftime('%Y%m%d_%H%M%S')})"
        db.execute("INSERT INTO templates (template_name, description, version) VALUES (?,?,?)",
                   (name, body.get("description", ""), body.get("version", "1.0")))
        tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        for p in body.get("phases", []):
            db.execute("INSERT INTO template_phases (template_id, phase_number, phase_name, duration_days) VALUES (?,?,?,?)",
                       (tid, p["phase_number"], p["phase_name"], p.get("duration_days", 14)))
            tp_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            for t in p.get("tasks", []):
                acfg = json.dumps(t.get("automation_config", {}))
                db.execute("""INSERT INTO template_tasks (template_phase_id, task_order, task_name, category,
                              description, ownership, instructions, estimated_minutes, priority,
                              automation_type, automation_config) VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                           (tp_id, t["task_order"], t["task_name"], t.get("category", "General"),
                            t.get("description", ""), t.get("ownership", "Platform Team"),
                            t.get("instructions", ""), t.get("estimated_minutes", 30),
                            t.get("priority", "medium"), t.get("automation_type", "manual"), acfg))
                tt_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
                for a in t.get("actions", []):
                    db.execute("INSERT INTO template_task_actions VALUES (NULL,?,?,?,?)",
                               (tt_id, a["action_type"], a["action_label"], a["action_url"]))
        for d in body.get("dependencies", []):
            db.execute("""INSERT INTO template_phase_dependencies
                          (template_id, phase_number, depends_on_phase_number) VALUES (?,?,?)""",
                       (tid, d["phase_number"], d["depends_on_phase_number"]))
        db.commit()
    return {"ok": True, "template_id": tid}


@app.post("/api/template/import-file")
async def api_import_template_file(file: UploadFile = File(...)):
    """Import template from Excel (.xlsx) or CSV file."""
    filename = file.filename.lower()
    content = await file.read()

    phases_data = []  # [{phase_number, phase_name, duration_days, tasks: [...]}]
    dependencies = []  # [{phase_number, depends_on_phase_number}]
    template_name = "Imported Template"

    if filename.endswith('.xlsx'):
        # Excel: multi-sheet format
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content))

            # Sheet 1: Phases
            if 'Phases' in wb.sheetnames:
                ws = wb['Phases']
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    r = dict(zip(headers, row))
                    phases_data.append({
                        "phase_number": int(r.get("phase_number", 1)),
                        "phase_name": str(r.get("phase_name", "Phase")),
                        "duration_days": int(r.get("duration_days", 14)),
                        "tasks": []
                    })

            # Sheet 2: Tasks
            if 'Tasks' in wb.sheetnames:
                ws = wb['Tasks']
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    r = dict(zip(headers, row))
                    pnum = int(r.get("phase_number", 1))
                    for p in phases_data:
                        if p["phase_number"] == pnum:
                            p["tasks"].append({
                                "task_order": int(r.get("task_order", 1)),
                                "task_name": str(r.get("task_name", "Task")),
                                "category": str(r.get("category", "General")),
                                "ownership": str(r.get("ownership", "Platform Team")),
                                "priority": str(r.get("priority", "medium")),
                                "estimated_minutes": int(r.get("estimated_minutes", 30)),
                                "automation_type": str(r.get("automation_type", "manual"))
                            })
                            break

            # Sheet 3: Dependencies
            if 'Dependencies' in wb.sheetnames:
                ws = wb['Dependencies']
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    r = dict(zip(headers, row))
                    dependencies.append({
                        "phase_number": int(r.get("phase_number", 1)),
                        "depends_on_phase_number": int(r.get("depends_on_phase_number", 0))
                    })

            # Template name from Info sheet or filename
            if 'Info' in wb.sheetnames:
                ws = wb['Info']
                for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                    if row[0] and str(row[0]).lower() == 'template_name' and row[1]:
                        template_name = str(row[1])
                        break
            else:
                template_name = file.filename.rsplit('.', 1)[0]

        except Exception as e:
            raise HTTPException(400, f"Excel parse error: {str(e)}")

    elif filename.endswith('.csv'):
        # CSV: flat format (phase info repeated per task)
        try:
            text = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            seen_phases = {}
            for row in reader:
                pnum = int(row.get("phase_number", 1))
                if pnum not in seen_phases:
                    seen_phases[pnum] = {
                        "phase_number": pnum,
                        "phase_name": row.get("phase_name", f"Phase {pnum}"),
                        "duration_days": int(row.get("duration_days", 14)),
                        "tasks": []
                    }
                seen_phases[pnum]["tasks"].append({
                    "task_order": int(row.get("task_order", len(seen_phases[pnum]["tasks"]) + 1)),
                    "task_name": row.get("task_name", "Task"),
                    "category": row.get("category", "General"),
                    "ownership": row.get("ownership", "Platform Team"),
                    "priority": row.get("priority", "medium"),
                    "estimated_minutes": int(row.get("estimated_minutes", 30)),
                    "automation_type": row.get("automation_type", "manual")
                })
            phases_data = list(seen_phases.values())
            template_name = file.filename.rsplit('.', 1)[0]
        except Exception as e:
            raise HTTPException(400, f"CSV parse error: {str(e)}")
    else:
        raise HTTPException(400, "Unsupported file type. Use .xlsx or .csv")

    if not phases_data:
        raise HTTPException(400, "No phases found in file")

    # Insert into DB (reuse JSON import logic)
    with get_db() as db:
        if db.execute("SELECT id FROM templates WHERE template_name=?", (template_name,)).fetchone():
            template_name += f" ({datetime.now().strftime('%Y%m%d_%H%M%S')})"
        db.execute("INSERT INTO templates (template_name, description, version) VALUES (?,?,?)",
                   (template_name, f"Imported from {file.filename}", "1.0"))
        tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        for p in phases_data:
            db.execute("INSERT INTO template_phases (template_id, phase_number, phase_name, duration_days) VALUES (?,?,?,?)",
                       (tid, p["phase_number"], p["phase_name"], p.get("duration_days", 14)))
            tp_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            for t in p.get("tasks", []):
                db.execute("""INSERT INTO template_tasks (template_phase_id, task_order, task_name, category,
                              description, ownership, instructions, estimated_minutes, priority,
                              automation_type, automation_config) VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                           (tp_id, t["task_order"], t["task_name"], t.get("category", "General"),
                            "", t.get("ownership", "Platform Team"), "",
                            t.get("estimated_minutes", 30), t.get("priority", "medium"),
                            t.get("automation_type", "manual"), "{}"))
        for d in dependencies:
            db.execute("""INSERT INTO template_phase_dependencies
                          (template_id, phase_number, depends_on_phase_number) VALUES (?,?,?)""",
                       (tid, d["phase_number"], d["depends_on_phase_number"]))
        db.commit()

    return {"ok": True, "template_id": tid, "template_name": template_name,
            "phases": len(phases_data), "tasks": sum(len(p["tasks"]) for p in phases_data)}


@app.get("/api/template/sample-excel")
async def api_sample_excel():
    """Download sample Excel template for import."""
    from openpyxl import Workbook
    wb = Workbook()

    # Info sheet
    ws_info = wb.active
    ws_info.title = "Info"
    ws_info.append(["template_name", "My Template Name"])
    ws_info.append(["description", "Template description here"])
    ws_info.append(["version", "1.0"])

    # Phases sheet
    ws_phases = wb.create_sheet("Phases")
    ws_phases.append(["phase_number", "phase_name", "duration_days"])
    ws_phases.append([1, "Prerequisites", 14])
    ws_phases.append([2, "Infrastructure Setup", 21])
    ws_phases.append([3, "Application Deployment", 14])

    # Tasks sheet
    ws_tasks = wb.create_sheet("Tasks")
    ws_tasks.append(["phase_number", "task_order", "task_name", "category", "ownership", "priority", "estimated_minutes", "automation_type"])
    ws_tasks.append([1, 1, "Request AWS Account", "Infrastructure", "Platform Team", "high", 30, "manual"])
    ws_tasks.append([1, 2, "Setup IAM Roles", "Infrastructure", "Platform Team", "high", 45, "semi_auto"])
    ws_tasks.append([2, 1, "Provision EKS Cluster", "Infrastructure", "Platform Team", "critical", 60, "auto"])
    ws_tasks.append([3, 1, "Deploy Application", "Deployment", "App Team", "high", 30, "semi_auto"])

    # Dependencies sheet
    ws_deps = wb.create_sheet("Dependencies")
    ws_deps.append(["phase_number", "depends_on_phase_number"])
    ws_deps.append([2, 1])
    ws_deps.append([3, 2])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_sample.xlsx"}
    )


@app.get("/api/template/sample-csv")
async def api_sample_csv():
    """Download sample CSV template for import."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["phase_number", "phase_name", "duration_days", "task_order", "task_name", "category", "ownership", "priority", "estimated_minutes", "automation_type"])
    writer.writerow([1, "Prerequisites", 14, 1, "Request AWS Account", "Infrastructure", "Platform Team", "high", 30, "manual"])
    writer.writerow([1, "Prerequisites", 14, 2, "Setup IAM Roles", "Infrastructure", "Platform Team", "high", 45, "semi_auto"])
    writer.writerow([2, "Infrastructure Setup", 21, 1, "Provision EKS Cluster", "Infrastructure", "Platform Team", "critical", 60, "auto"])
    writer.writerow([3, "Application Deployment", 14, 1, "Deploy Application", "Deployment", "App Team", "high", 30, "semi_auto"])

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template_sample.csv"}
    )


# ── Template Phase CRUD ────────────────────────────────────────────────────

@app.post("/api/template/{template_id}/phase")
async def api_create_template_phase(template_id: int, request: Request):
    body = await request.json()
    with get_db() as db:
        db.execute("""INSERT INTO template_phases (template_id, phase_number, phase_name, duration_days)
                      VALUES (?,?,?,?)""",
                   (template_id, body["phase_number"], body["phase_name"], body.get("duration_days", 14)))
        db.commit()
        pid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return {"ok": True, "phase_id": pid}


@app.put("/api/template/phase/{phase_id}")
async def api_update_template_phase(phase_id: int, request: Request):
    body = await request.json()
    with get_db() as db:
        db.execute("UPDATE template_phases SET phase_name=?, phase_number=?, duration_days=? WHERE id=?",
                   (body["phase_name"], body["phase_number"], body.get("duration_days", 14), phase_id))
        db.commit()
    return {"ok": True}


@app.delete("/api/template/phase/{phase_id}")
async def api_delete_template_phase(phase_id: int):
    with get_db() as db:
        db.execute("DELETE FROM template_phases WHERE id=?", (phase_id,))
        db.commit()
    return {"ok": True}


# ── Template Task CRUD ─────────────────────────────────────────────────────

@app.post("/api/template/phase/{phase_id}/task")
async def api_create_template_task(phase_id: int, request: Request):
    body = await request.json()
    with get_db() as db:
        mx = db.execute("SELECT COALESCE(MAX(task_order),0) FROM template_tasks WHERE template_phase_id=?",
                        (phase_id,)).fetchone()[0]
        acfg = json.dumps(body.get("automation_config", {}))
        db.execute("""INSERT INTO template_tasks (template_phase_id, task_order, task_name, category,
                      description, ownership, instructions, jira_reference, sample_reference, notes,
                      automation_type, automation_config)
                      VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                   (phase_id, mx + 1, body["task_name"], body.get("category", "General"),
                    body.get("description", ""), body.get("ownership", "Platform Team"),
                    body.get("instructions", ""), body.get("jira_reference"),
                    body.get("sample_reference"), body.get("notes"),
                    body.get("automation_type", "manual"), acfg))
        db.commit()
        tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    return {"ok": True, "task_id": tid}


@app.put("/api/template/task/{task_id}")
async def api_update_template_task(task_id: int, request: Request):
    body = await request.json()
    acfg = json.dumps(body.get("automation_config", {}))
    with get_db() as db:
        db.execute("""UPDATE template_tasks SET task_name=?, category=?, description=?,
                      ownership=?, instructions=?, jira_reference=?, sample_reference=?, notes=?,
                      automation_type=?, automation_config=?
                      WHERE id=?""",
                   (body["task_name"], body.get("category", ""), body.get("description", ""),
                    body.get("ownership", "Platform Team"), body.get("instructions", ""),
                    body.get("jira_reference"), body.get("sample_reference"), body.get("notes"),
                    body.get("automation_type", "manual"), acfg, task_id))
        db.commit()
    return {"ok": True}


@app.delete("/api/template/task/{task_id}")
async def api_delete_template_task(task_id: int):
    with get_db() as db:
        db.execute("DELETE FROM template_tasks WHERE id=?", (task_id,))
        db.commit()
    return {"ok": True}


# ── Template Dependencies ──────────────────────────────────────────────────

@app.post("/api/template/{template_id}/dependency")
async def api_add_template_dep(template_id: int, request: Request):
    body = await request.json()
    with get_db() as db:
        db.execute("""INSERT OR IGNORE INTO template_phase_dependencies
                      (template_id, phase_number, depends_on_phase_number) VALUES (?,?,?)""",
                   (template_id, body["phase_number"], body["depends_on_phase_number"]))
        db.commit()
    return {"ok": True}


@app.delete("/api/template/dependency/{dep_id}")
async def api_remove_template_dep(dep_id: int):
    with get_db() as db:
        db.execute("DELETE FROM template_phase_dependencies WHERE id=?", (dep_id,))
        db.commit()
    return {"ok": True}


@app.get("/api/template/{template_id}/dag")
async def api_template_dag(template_id: int):
    with get_db() as db:
        t_phases = rows(db.execute("""
            SELECT tp.*,
                (SELECT COUNT(*) FROM template_tasks WHERE template_phase_id=tp.id) as task_count
            FROM template_phases tp WHERE tp.template_id=? ORDER BY tp.phase_number
        """, (template_id,)).fetchall())
        t_deps = rows(db.execute("""
            SELECT depends_on_phase_number as from_phase, phase_number as to_phase
            FROM template_phase_dependencies WHERE template_id=?
        """, (template_id,)).fetchall())

    nodes = [{
        "id": p["phase_number"],
        "label": f"P{p['phase_number']}: {p['phase_name']}\n{p['task_count']} tasks | {p['duration_days']}d",
        "color": {"background": "#3b82f6", "border": "#2563eb",
                  "highlight": {"background": "#60a5fa", "border": "#1e293b"}},
        "font": {"color": "#ffffff", "size": 13, "face": "Inter, system-ui, sans-serif"},
        "shape": "box", "borderWidth": 2, "shadow": True,
        "margin": {"top": 12, "bottom": 12, "left": 16, "right": 16},
        "widthConstraint": {"minimum": 160, "maximum": 200},
    } for p in t_phases]

    # Edges flow from prerequisite TO dependent phase
    edges = [{
        "from": d["from_phase"], "to": d["to_phase"],
        "arrows": {"to": {"enabled": True, "scaleFactor": 0.7, "type": "arrow"}},
        "color": {"color": "#94a3b8", "highlight": "#3b82f6"},
        "width": 2,
        "smooth": {"type": "curvedCW", "roundness": 0.15},
    } for d in t_deps]

    return {"nodes": nodes, "edges": edges}


# ── Project creation from template ─────────────────────────────────────────

@app.post("/api/project/create")
async def api_create_project(request: Request):
    body = await request.json()
    template_id = body["template_id"]
    start_str = body["start_date"]
    start_date = datetime.strptime(start_str, "%Y-%m-%d")

    with get_db() as db:
        total_dur = db.execute(
            "SELECT COALESCE(SUM(duration_days), 90) FROM template_phases WHERE template_id=?",
            (template_id,)
        ).fetchone()[0]
        target = body.get("target_completion") or (start_date + timedelta(days=total_dur)).strftime("%Y-%m-%d")

        db.execute("""INSERT INTO projects (project_name, domain, environment, team_name, team_email,
                      jira_project, aws_account_id, aws_region, eks_cluster_name,
                      status, start_date, target_completion, template_id)
                      VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                   (body["project_name"], body["domain"], body["environment"],
                    body.get("team_name", ""), body.get("team_email", ""),
                    body.get("jira_project", ""), body.get("aws_account_id", ""),
                    body.get("aws_region", "us-east-1"), body.get("eks_cluster_name", ""),
                    "in_progress", start_str, target, template_id))
        project_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        create_project_from_template(db, project_id, template_id, start_date)
        db.commit()
    return {"ok": True, "project_id": project_id}


@app.put("/api/project/{project_id}")
async def api_update_project(project_id: int, request: Request):
    """Update project metadata."""
    body = await request.json()
    with get_db() as db:
        db.execute("""UPDATE projects SET
                      project_name=COALESCE(?, project_name),
                      domain=COALESCE(?, domain),
                      environment=COALESCE(?, environment),
                      team_name=COALESCE(?, team_name),
                      team_email=COALESCE(?, team_email),
                      jira_project=COALESCE(?, jira_project),
                      aws_account_id=COALESCE(?, aws_account_id),
                      aws_region=COALESCE(?, aws_region),
                      eks_cluster_name=COALESCE(?, eks_cluster_name),
                      start_date=COALESCE(?, start_date),
                      target_completion=COALESCE(?, target_completion),
                      status=COALESCE(?, status)
                      WHERE id=?""",
                   (body.get("project_name"), body.get("domain"), body.get("environment"),
                    body.get("team_name"), body.get("team_email"), body.get("jira_project"),
                    body.get("aws_account_id"), body.get("aws_region"), body.get("eks_cluster_name"),
                    body.get("start_date"), body.get("target_completion"), body.get("status"),
                    project_id))
        db.commit()
    return {"ok": True}


@app.delete("/api/project/{project_id}")
async def api_delete_project(project_id: int):
    """Delete a project and all its phases/tasks."""
    with get_db() as db:
        db.execute("DELETE FROM projects WHERE id=?", (project_id,))
        db.commit()
    return {"ok": True}


@app.get("/api/domains")
async def api_list_domains():
    with get_db() as db:
        return rows(db.execute("SELECT * FROM domains WHERE is_active=1 ORDER BY code").fetchall())


# ── Deadline Tracking ───────────────────────────────────────────────────────

def calc_variance(agreed_date: str, actual_date: str) -> int:
    """Calculate variance in days (positive = late, negative = early)."""
    if not agreed_date or not actual_date:
        return None
    agreed = datetime.strptime(agreed_date, "%Y-%m-%d").date()
    actual = datetime.strptime(actual_date, "%Y-%m-%d").date()
    return (actual - agreed).days


@app.post("/api/phase/{phase_id}/deadline")
async def api_set_phase_deadline(phase_id: int, request: Request):
    """Set or update deadline for a phase by ownership (Platform Team / App Team)."""
    body = await request.json()
    ownership = body.get("ownership", "Platform Team")
    planned = body.get("planned_date")
    agreed = body.get("agreed_date")
    actual = body.get("actual_date")
    notes = body.get("notes", "")
    variance = calc_variance(agreed, actual)

    with get_db() as db:
        # Upsert: update if exists, insert if not
        existing = db.execute(
            "SELECT id FROM phase_deadlines WHERE phase_id=? AND ownership=?",
            (phase_id, ownership)
        ).fetchone()
        if existing:
            db.execute("""UPDATE phase_deadlines SET planned_date=?, agreed_date=?, actual_date=?,
                          variance_days=?, notes=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                       (planned, agreed, actual, variance, notes, existing[0]))
        else:
            db.execute("""INSERT INTO phase_deadlines (phase_id, ownership, planned_date, agreed_date,
                          actual_date, variance_days, notes) VALUES (?,?,?,?,?,?,?)""",
                       (phase_id, ownership, planned, agreed, actual, variance, notes))
        db.commit()
    return {"ok": True, "variance_days": variance}


@app.get("/api/phase/{phase_id}/deadlines")
async def api_get_phase_deadlines(phase_id: int):
    """Get all deadlines for a phase."""
    with get_db() as db:
        return rows(db.execute(
            "SELECT * FROM phase_deadlines WHERE phase_id=? ORDER BY ownership",
            (phase_id,)
        ).fetchall())


@app.post("/api/task/{task_id}/deadline")
async def api_set_task_deadline(task_id: int, request: Request):
    """Set or update deadline for a critical task."""
    body = await request.json()
    planned = body.get("planned_date")
    agreed = body.get("agreed_date")
    actual = body.get("actual_date")
    notes = body.get("notes", "")
    variance = calc_variance(agreed, actual)

    with get_db() as db:
        existing = db.execute("SELECT id FROM task_deadlines WHERE task_id=?", (task_id,)).fetchone()
        if existing:
            db.execute("""UPDATE task_deadlines SET planned_date=?, agreed_date=?, actual_date=?,
                          variance_days=?, notes=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                       (planned, agreed, actual, variance, notes, existing[0]))
        else:
            db.execute("""INSERT INTO task_deadlines (task_id, planned_date, agreed_date,
                          actual_date, variance_days, notes) VALUES (?,?,?,?,?,?)""",
                       (task_id, planned, agreed, actual, variance, notes))
        db.commit()
    return {"ok": True, "variance_days": variance}


@app.get("/api/project/{project_id}/variance")
async def api_project_variance(project_id: int):
    """Get deadline variance report for a project."""
    with get_db() as db:
        # Phase-level deadlines
        phase_deadlines = rows(db.execute("""
            SELECT ph.phase_number, ph.phase_name, pd.ownership, pd.planned_date, pd.agreed_date,
                   pd.actual_date, pd.variance_days, pd.notes
            FROM phases ph
            LEFT JOIN phase_deadlines pd ON pd.phase_id = ph.id
            WHERE ph.project_id = ?
            ORDER BY ph.phase_number, pd.ownership
        """, (project_id,)).fetchall())

        # Task-level deadlines (critical tasks only)
        task_deadlines = rows(db.execute("""
            SELECT t.task_name, t.ownership, ph.phase_number, td.planned_date, td.agreed_date,
                   td.actual_date, td.variance_days, td.notes
            FROM tasks t
            JOIN phases ph ON t.phase_id = ph.id
            JOIN task_deadlines td ON td.task_id = t.id
            WHERE ph.project_id = ?
            ORDER BY ph.phase_number, t.task_order
        """, (project_id,)).fetchall())

        # Summary stats
        stats = db.execute("""
            SELECT
                COUNT(*) as total_deadlines,
                SUM(CASE WHEN variance_days IS NOT NULL AND variance_days <= 0 THEN 1 ELSE 0 END) as on_time,
                SUM(CASE WHEN variance_days > 0 AND variance_days <= 3 THEN 1 ELSE 0 END) as minor_slip,
                SUM(CASE WHEN variance_days > 3 THEN 1 ELSE 0 END) as major_slip,
                AVG(variance_days) as avg_variance
            FROM phase_deadlines pd
            JOIN phases ph ON pd.phase_id = ph.id
            WHERE ph.project_id = ?
        """, (project_id,)).fetchone()

    return {
        "phase_deadlines": phase_deadlines,
        "task_deadlines": task_deadlines,
        "stats": {
            "total_deadlines": stats[0] or 0,
            "on_time": stats[1] or 0,
            "minor_slip": stats[2] or 0,
            "major_slip": stats[3] or 0,
            "avg_variance": round(stats[4], 1) if stats[4] else 0
        }
    }


# ── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    db_path = os.path.join(BASE_DIR, DB_PATH)
    if not os.path.exists(db_path):
        print("Initializing database...")
        os.chdir(BASE_DIR)
        init_database()
    print("Starting Application Onboarding Portal at http://localhost:8000")
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
